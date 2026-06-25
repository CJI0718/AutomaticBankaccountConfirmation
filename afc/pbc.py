"""회사제시(PBC) 예금명세 로더 + 조회서 ↔ PBC 양방향 대사 엔진.

감사조서 4200 '조회서대사'의 핵심: 회사가 준 명세(PBC)와 은행 조회서를 맞춰
  (1) 금액 차이,
  (2) 회사명세엔 있는데 조회서가 안 온 계좌 (PBC only — 완전성 위험),
  (3) 조회서엔 있는데 회사명세에 없는 계좌 (조회서 only — 누락 위험)
를 잡는다. 두 방향 모두 확인하는 것이 완전성(Completeness) 절차의 요점.

PBC 양식은 클라이언트마다 다르므로 헤더 행을 탐지해 매핑한다(옵티팜 계정명세서의
보통예금/기타제예금/정기예금 명세 양식 기준, 변형 허용).
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from afc.reconciliation import ReconRow

# PBC 시트 헤더 컬럼 별칭 → 표준 키.
_HEADER_ALIASES = {
    "금융기관": "institution", "은행": "institution", "기관": "institution",
    "계좌번호": "account", "계좌": "account",
    "금액": "amount", "잔액": "amount", "원화금액": "amount", "환산후": "amount",
    "적요": "memo", "상품": "memo", "종류": "memo", "비고": "note",
    "사업장": "site", "통화": "currency",
}
# 데이터가 아닌 집계/대조 행.
_STOP_TOKENS = ("합계", "합 계", "소계", "대차대조표", "비교", "TB", "B/S")
_DEPOSIT_SHEET_HINT = ("예금", "명세", "현금", "금융상품")


def normalize_account(value: object) -> str:
    """계좌번호에서 숫자만 남긴다. '087-17-022205' → '08717022205'."""
    return re.sub(r"\D", "", str(value)) if value is not None else ""


@dataclass(frozen=True)
class PbcRow:
    institution: str
    account_raw: str
    account_key: str
    amount: float | None
    currency: str | None
    memo: str | None
    source_sheet: str


def _to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(re.sub(r"[^\d.\-]", "", str(value)))
    except ValueError:
        return None


def _find_header(ws) -> tuple[int, dict[int, str]] | None:
    """헤더 행(번호)과 {열번호: 표준키} 매핑을 찾는다. 계좌+금액 필수."""
    for r in range(1, min(ws.max_row, 12) + 1):
        mapping: dict[int, str] = {}
        for c in range(1, min(ws.max_column, 15) + 1):
            label = ws.cell(r, c).value
            if not isinstance(label, str):
                continue
            key = _HEADER_ALIASES.get(label.replace(" ", ""))
            if key:
                mapping[c] = key
        if "account" in mapping.values() and "amount" in mapping.values():
            return r, mapping
    return None


def load_pbc(path: Path) -> list[PbcRow]:
    """PBC 워크북에서 예금/금융상품 명세 행을 추출한다."""
    wb = load_workbook(path, data_only=True)
    rows: list[PbcRow] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        # 시트명 힌트 또는 헤더 존재로 예금명세 판별.
        header = _find_header(ws)
        if header is None:
            continue
        name_ok = any(h in sn for h in _DEPOSIT_SHEET_HINT)
        if not name_ok:
            # 시트명이 애매해도 헤더가 명확하면 채택.
            pass
        hdr_row, colmap = header
        for r in range(hdr_row + 1, ws.max_row + 1):
            cells = {key: ws.cell(r, c).value for c, key in colmap.items()}
            acct = cells.get("account")
            if acct is None or str(acct).strip() == "":
                continue
            joined = " ".join(str(v) for v in cells.values() if v is not None)
            if any(tok in joined for tok in _STOP_TOKENS):
                continue
            key = normalize_account(acct)
            if len(key) < 6:  # 계좌번호로 보기 어려움 (집계/메모 행)
                continue
            memo = cells.get("memo")
            currency = cells.get("currency")
            if not currency and memo and "외화" in str(memo):
                currency = None  # 외화지만 통화 미상 → 추후 보정
            rows.append(
                PbcRow(
                    institution=str(cells.get("institution") or "").strip(),
                    account_raw=str(acct).strip(),
                    account_key=key,
                    amount=_to_float(cells.get("amount")),
                    currency=str(currency).strip() if currency else "KRW",
                    memo=str(memo).strip() if memo else None,
                    source_sheet=sn,
                )
            )
    return rows


# ── 대사 ────────────────────────────────────────────────────────────────────

def _norm_currency(currency: str | None) -> str:
    """통화 코드 정규화. None/WON → KRW (원화)."""
    if not currency or currency in ("WON",):
        return "KRW"
    return currency


@dataclass(frozen=True)
class MatchedAccount:
    account_key: str
    institution: str
    category: str
    currency: str
    confirm_amount: float        # 조회서 (계좌·통화 단위)
    pbc_amount: float            # PBC 금액

    @property
    def diff(self) -> float:
        return round(self.confirm_amount - self.pbc_amount, 2)

    @property
    def status(self) -> str:
        return "일치" if abs(self.diff) < 0.5 else "차이"

    @property
    def currencies(self) -> tuple[str, ...]:  # 시트 표시 호환
        return (self.currency,)


@dataclass
class ReconResult:
    matched: list[MatchedAccount] = field(default_factory=list)
    confirm_only: list[tuple[str, str, str, float]] = field(default_factory=list)  # key, inst, cat, amt
    pbc_only: list[PbcRow] = field(default_factory=list)

    def summary(self) -> dict[str, int]:
        return {
            "matched": len(self.matched),
            "diff": sum(1 for m in self.matched if m.status == "차이"),
            "confirm_only": len(self.confirm_only),
            "pbc_only": len(self.pbc_only),
        }


def _confirm_by_account(recon_rows: list[ReconRow]) -> dict[tuple[str, str], dict]:
    """조회서 예금(자산) 행을 (계좌, 통화) 단위로 집계. 통화별 잔액은 합산 불가하므로 분리.
    PBC(회사 예금명세)는 금융자산 명세이므로 부채(차입금)는 대사 대상에서 제외한다."""
    agg: dict[tuple[str, str], dict] = {}
    for row in recon_rows:
        if row.side != "자산":
            continue
        key = (normalize_account(row.account_no), _norm_currency(row.currency))
        slot = agg.setdefault(
            key, {"institution": row.institution, "category": row.category, "amount": 0.0}
        )
        slot["amount"] += row.amount or 0.0
    return agg


def reconcile(recon_rows: list[ReconRow], pbc_rows: list[PbcRow]) -> ReconResult:
    """조회서 ↔ PBC 양방향 대사. 키 = (계좌번호, 통화)."""
    confirm = _confirm_by_account(recon_rows)
    pbc = {(p.account_key, _norm_currency(p.currency)): p for p in pbc_rows}
    result = ReconResult()

    for key, c in confirm.items():
        acct, currency = key
        if key in pbc:
            p = pbc[key]
            result.matched.append(
                MatchedAccount(
                    account_key=acct,
                    institution=c["institution"] or p.institution,
                    category=c["category"],
                    currency=currency,
                    confirm_amount=round(c["amount"], 2),
                    pbc_amount=round(p.amount or 0.0, 2),
                )
            )
        else:
            result.confirm_only.append((acct, c["institution"], c["category"], round(c["amount"], 2)))

    for key, p in pbc.items():
        if key not in confirm:
            result.pbc_only.append(p)

    result.matched.sort(key=lambda m: (m.status != "차이", m.institution, m.account_key))
    return result
