"""삼화전기 조회서에서 '회사제시(PBC) 예금명세' 모의 파일을 생성한다 (데모용).

실제로는 회사가 PBC를 주지만, 삼화전기 표본엔 PBC가 없어 업무 재현을 위해
조회서 추출 결과를 옵티팜 계정명세서 양식으로 옮기되, 감사에서 실제로 마주치는
불일치를 의도적으로 심는다:
  · 일부 계좌 잔액 변경      -> 금액 차이
  · 일부 계좌 누락           -> 조회서에만 존재(회사명세 누락, 완전성 위험)
  · 가공 계좌 1건 추가        -> 회사명세에만 존재(조회서 미회수, 완전성 위험)

사용: python -m afc.make_sample_pbc "<삼화전기 zip>" -o "<출력 xlsx>"
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook

from afc.institutions import load_institution_config
from afc.run import extract_records

# 결정론적 시딩 — 데모 재현성을 위해 계좌번호로 고정 지정.
_AMOUNT_TWEAKS = {  # 계좌(끝부분) : 차이를 만들 변경 금액
    "140007161759": 2_410_000_000.0,   # 실제 2,414,000,000 → 차이 4,000,000
    "1005202813898": 225_000_000.0,    # 실제 225,423,323 → 차이 423,323
}
_DROP_ACCOUNTS = {"140011076230", "100029387386"}  # PBC에서 제외 → 명세누락
_PHANTOM = ("우리은행", "정기예금", "1002-999-888777", "KRW", 500_000_000.0, "회사명세 가공계좌(데모)")


def build_pbc_rows(zip_path: Path):
    config = load_institution_config()
    records = extract_records(zip_path, config)
    agg: dict[tuple[str, str], dict] = {}
    for rec in records:
        for d in rec.deposits:
            currency = d.balance.currency or "KRW"
            slot = agg.setdefault((d.account_no, currency),
                                  {"inst": d.header.institution_name,
                                   "memo": d.product_type, "amt": 0.0})
            slot["amt"] += d.balance.amount or 0.0
    rows = []
    for (acct, currency), v in agg.items():
        if acct in _DROP_ACCOUNTS:
            continue
        amount = _AMOUNT_TWEAKS.get(acct, round(v["amt"], 2))
        rows.append((v["inst"], v["memo"], acct, currency, amount, ""))
    rows.sort(key=lambda r: (r[0], r[2], r[3]))
    rows.append(_PHANTOM)
    return rows


def write_pbc(rows, path: Path, client: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "보통예금"
    ws.cell(2, 1, "보 통 예 금 명 세 서")
    ws.cell(4, 1, f"{client} (회사제시 PBC — 데모 생성)")
    headers = ["금융기관", "적요", "계좌번호", "통화", "금액", "비고"]
    for j, h in enumerate(headers, start=1):
        ws.cell(5, j, h)
    r = 6
    for inst, memo, acct, currency, amount, note in rows:
        ws.cell(r, 1, inst); ws.cell(r, 2, memo); ws.cell(r, 3, acct)
        ws.cell(r, 4, currency)
        ws.cell(r, 5, amount).number_format = "#,##0.00"
        ws.cell(r, 6, note)
        r += 1
    ws.cell(r, 1, "합 계")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="삼화전기 조회서 → 모의 PBC 예금명세 생성")
    parser.add_argument("zip_path", type=Path)
    parser.add_argument("-o", "--output", type=Path, required=True)
    parser.add_argument("--client", default="삼화전기")
    args = parser.parse_args()
    rows = build_pbc_rows(args.zip_path)
    write_pbc(rows, args.output, args.client)
    print(f"[모의 PBC 생성] {args.output} — {len(rows)}계좌 "
          f"(차이 {len(_AMOUNT_TWEAKS)}, 명세누락 {len(_DROP_ACCOUNTS)}, 가공 1)")


if __name__ == "__main__":
    main()
