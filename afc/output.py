"""다중 파일 출력 — 1섹션=1시트, 카테고리별 파일 + 요약 + PBC대사.

한 시트에 여러 표를 쌓으면 칼럼 수가 제각각이라 가독성이 무너진다. find_tables가
표를 개별 단위로 주므로, 표 하나를 시트 하나로 깔끔히 분리하고 파일도 용도별로 나눈다:

  금융기관조회서_BANK_<client>.xlsx        0_요약(라이브 수식) + 섹션별 시트
  금융기관조회서_INSURANCE_<client>.xlsx
  금융기관조회서_GUARANTEE_<client>.xlsx
  금융기관조회서_INVESTMENT_<client>.xlsx
  금융기관조회서_PBC대사_<client>.xlsx       대사·계정과목별·금융기관별·통화별·PBC대사

각 카테고리 파일의 첫 시트 '0_요약'은 그 파일 내부 데이터 시트를 SUMIFS/COUNTIFS
라이브 수식으로 집계한다(별도 요약 파일을 두지 않음 — 같은 워크북 참조라 견고).
"""
from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from afc.excel_writer import (
    _autosize,
    _money_fmt,
    _style_header,
    build_pbc_recon_sheet,
    build_recon_sheet,
    _write_summary,
)
from afc.extract import ConfirmationRecord, format_date8, parse_money
from afc.institutions import InstitutionConfig
from afc.reconciliation import build_recon_rows, load_account_mapping
from afc.sections import ID_COLUMNS, load_section_spec

# 0_요약 시트 스타일 (조 회계사 양식).
_NAVY = PatternFill("solid", fgColor="1F4E78")
_LIGHTBLUE = PatternFill("solid", fgColor="D9E1F2")
_GRAY = PatternFill("solid", fgColor="F2F2F2")
_ACCT = "#,##0"
_CIRCLED = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭"
_INST_COL = get_column_letter(ID_COLUMNS.index("금융기관명") + 1)  # 금융기관명 = D열

# 금액성 칼럼 키워드. 율/순위/좌수 등은 이 키워드를 포함하지 않으므로 자연히 제외된다
# (예: '설정순위'엔 '금액'이 없고, '선순위 설정금액'엔 '금액'이 있어 정확히 갈린다).
_MONEY_KW = (
    "금액", "한도", "잔액", "잔고", "부보", "보장", "보험료", "보험가입",
    "적립금", "환급금", "보증금", "예수금", "평가액", "평가금", "출자금", "감정", "설정금",
)
# (카테고리, 섹션 id) → 짧은 시트명 접미.
_SECTION_SHORT = {
    "BANK": {
        "1": "예금", "2-1": "총한도", "2-2": "대출", "3": "지급보증", "4": "파생",
        "5": "담보연대보증", "6-1": "전자어음", "6-2": "수표어음", "7-1": "미발행어음",
        "7-2": "미결제어음", "7-3": "수표어음2", "8": "담보견질", "9": "담보제공", "10": "당좌거래",
    },
    "INSURANCE": {
        "1": "보험거래", "2": "대출", "3": "지급보증", "4": "담보견질",
        "5": "담보연대보증", "6": "담보제공",
    },
    "GUARANTEE": {
        "1": "보증", "2": "신용보험", "3": "기타공제보험", "4": "대출융자투자",
        "5": "출자금", "6": "한도거래", "7": "담보어음수표", "8": "담보연대보증", "9": "담보제공",
    },
    "INVESTMENT": {
        "1": "유가증권", "2": "상세명세", "3": "대출", "4": "지급보증", "5": "파생",
        "6": "담보연대보증", "7": "담보견질", "8": "담보제공", "9": "거래내역",
    },
}
_CATEGORY_FILE = {"BANK": "BANK", "INSURANCE": "INSURANCE",
                  "GUARANTEE": "GUARANTEE", "INVESTMENT": "INVESTMENT"}


def _is_money_col(name: str) -> bool:
    flat = (name or "").replace(" ", "")   # '금    액' → '금액'
    return any(k in flat for k in _MONEY_KW)


def _fmt_cell(value: str) -> str:
    s = (value or "").strip()
    if s.isdigit() and len(s) == 8:  # 8자리 = 날짜
        return format_date8(s)
    return s


def _expand_columns(data_cols: tuple[str, ...]):
    """금액 칼럼은 통화/금액 두 칼럼으로 펼친다. (출력컬럼명들, [(kind, src_idx)…])."""
    out_cols: list[str] = []
    plan: list[tuple[str, int]] = []
    for i, name in enumerate(data_cols):
        if _is_money_col(name):
            out_cols += [f"{name}_통화", f"{name}_금액"]
            plan.append(("money", i))
        else:
            out_cols.append(name)
            plan.append(("plain", i))
    return out_cols, plan


def _safe_sheet_name(name: str) -> str:
    return re.sub(r'[\\/?*\[\]:]', "_", name)[:31]


def _field_label(name: str) -> str:
    """금액 칼럼명 → 짧은 라벨. '금액_약정한도액'→'한도', '금액_대출금액'→'잔액'."""
    n = name.replace(" ", "")
    for kw, label in (("약정한도", "한도"), ("대출금액", "잔액"), ("실행", "실행"),
                      ("선순위", "선순위"), ("한도", "한도"), ("설정금", "설정"), ("감정", "감정"),
                      ("부보", "부보"), ("보장", "보장"), ("환급금", "환급"), ("적립금", "적립")):
        if kw in n:
            return label
    return n.replace("_금액", "").replace("금액", "") or "금액"


def _money_fields(spec) -> list[tuple[str, str, str, int]]:
    """섹션의 금액 필드들 → (라벨, 통화칼럼letter, 금액칼럼letter, 원본열idx)."""
    fields: list[tuple[str, str, str, int]] = []
    base = len(ID_COLUMNS)
    out_pos = 0
    for src_idx, name in enumerate(spec.data_columns):
        if _is_money_col(name):
            cur_letter = get_column_letter(base + out_pos + 1)
            amt_letter = get_column_letter(base + out_pos + 2)
            fields.append((_field_label(name), cur_letter, amt_letter, src_idx))
            out_pos += 2
        else:
            out_pos += 1
    return fields


def _present_currencies(records, sid: str, src_idx: int, config) -> list[str]:
    """해당 섹션·금액필드에 실제 등장하는 통화 코드(등장순 정렬)."""
    found: set[str] = set()
    for rec in records:
        t = rec.section_tables.get(sid)
        if not t:
            continue
        for row in t.rows:
            if src_idx < len(row):
                m = parse_money(row[src_idx].replace(" ", ""), config)
                if m.amount is not None:            # 금액 있으면 빈 통화는 KRW로
                    found.add(m.currency or "KRW")
    order = ["KRW", "WON", "USD", "EUR", "JPY", "CNY", "HKD"]
    return sorted(found, key=lambda c: (order.index(c) if c in order else 99, c))


def _distinct_institutions(records) -> list[str]:
    return sorted({r.header.institution_name for r in records})


def build_credit_summary_sheet(ws, category: str, records, specs, config) -> None:
    """파일 내부 데이터 시트를 SUMIFS/COUNTIFS 라이브 수식으로 집계 (조 회계사 양식).

    금융기관 × (금액필드 × 통화) + 건수, 합계 행. 같은 워크북 내부 참조라 견고하다.
    """
    client = records[0].header.company_name if records else ""
    date = records[0].header.confirmation_date if records else ""
    insts = _distinct_institutions(records)

    ttl = ws.cell(1, 1, f"{client} 금융기관조회서 통합 요약")
    ttl.font = Font(bold=True, size=14, color="1F4E78")
    ws.cell(2, 1, f"조회기준일: {date}  ·  통화별 분리 집계 (환율 환산 없음)  ·  "
                  "수식 연동(데이터 수정 시 자동 갱신)").font = Font(size=10, color="595959")

    shorts = _SECTION_SHORT.get(category, {})
    r = 4
    n = 0
    for spec in specs:
        fields = _money_fields(spec)
        if not any(rec.section_tables.get(spec.id) and rec.section_tables[spec.id].rows for rec in records):
            continue
        if not fields:
            continue
        sheet = _safe_sheet_name(f"{spec.id}_{shorts.get(spec.id, spec.id)}")
        n += 1
        circ = _CIRCLED[n - 1] if n <= len(_CIRCLED) else f"({n})"
        # 섹션 제목 바
        labels = " / ".join(dict.fromkeys(f[0] for f in fields))
        bar = ws.cell(r, 1, f"{circ} {shorts.get(spec.id, spec.id)} ({sheet}) — {labels} 통화별 집계")
        bar.font = Font(bold=True, color="FFFFFF")
        bar.fill = _NAVY
        r += 1

        # 칼럼 구성: 금융기관 | (필드별 통화) … | 건수
        single = len(fields) == 1
        col_defs: list[tuple[str, str, str, str]] = []  # (헤더, 통화letter, 금액letter, 통화코드)
        for label, cur_l, amt_l, src_idx in fields:
            for cur in _present_currencies(records, spec.id, src_idx, config) or ["KRW"]:
                head = cur if single else f"{label} {cur}"
                col_defs.append((head, cur_l, amt_l, cur))
        headers = ["금융기관"] + [cd[0] for cd in col_defs] + ["건수"]
        for j, h in enumerate(headers, start=1):
            cell = ws.cell(r, j, h)
            cell.font = Font(bold=True)
            cell.fill = _LIGHTBLUE
            cell.alignment = Alignment(horizontal="center")
        hdr_row = r
        r += 1

        first = r
        for inst in insts:
            ws.cell(r, 1, inst)
            for j, (head, cur_l, amt_l, cur) in enumerate(col_defs, start=2):
                f = (f"=SUMIFS('{sheet}'!${amt_l}:${amt_l},'{sheet}'!${_INST_COL}:${_INST_COL},"
                     f'$A{r},\'{sheet}\'!${cur_l}:${cur_l},"{cur}")')
                c = ws.cell(r, j, f)
                c.number_format = _ACCT
            cnt = ws.cell(r, len(headers), f"=COUNTIFS('{sheet}'!${_INST_COL}:${_INST_COL},$A{r})")
            cnt.alignment = Alignment(horizontal="center")
            r += 1
        # 합계
        tot = ws.cell(r, 1, "합계")
        tot.font = Font(bold=True)
        tot.fill = _GRAY
        for j in range(2, len(headers) + 1):
            col = get_column_letter(j)
            c = ws.cell(r, j, f"=SUM({col}{first}:{col}{r-1})")
            c.font = Font(bold=True)
            c.fill = _GRAY
            c.number_format = _ACCT
            if j == len(headers):
                c.alignment = Alignment(horizontal="center")
        ws.cell(r, 1).fill = _GRAY
        r += 2

    ws.column_dimensions["A"].width = 22
    for j in range(2, 14):
        ws.column_dimensions[get_column_letter(j)].width = 15
    ws.freeze_panes = "B5"


def render_section_sheet(ws, section_id, spec, records, config: InstitutionConfig, client: str) -> None:
    out_cols, plan = _expand_columns(spec.data_columns)
    columns = ID_COLUMNS + out_cols
    ws.cell(1, 1, f"{section_id}. {spec.header}").font = Font(bold=True)
    for j, c in enumerate(columns, start=1):
        ws.cell(2, j, c)
    _style_header(ws, 2, len(columns))
    money_out_idx = {j for j, c in enumerate(columns) if c.endswith("_금액")}

    r = 3
    for rec in sorted(records, key=lambda x: (x.header.institution_name, x.header.business_no)):
        h = rec.header
        base = [None, h.company_name, h.business_no, h.institution_name, h.confirmation_date]
        table = rec.section_tables.get(section_id)
        rows = table.rows if table else []
        data_rows = rows if rows else [None]  # 해당없음도 1행
        for cells in data_rows:
            if cells is None:
                vals = list(base) + ["해당 없음"]
            else:
                vals = list(base)
                for kind, idx in plan:
                    cell = cells[idx] if idx < len(cells) else ""
                    if kind == "money":
                        m = parse_money(cell.replace(" ", ""), config)
                        # 국내 금액은 통화코드가 빠진 채 숫자만 오는 경우가 많다.
                        # 금액이 있는데 통화가 비면 KRW로 본다(0_요약 SUMIFS 집계 누락 방지).
                        cur = m.currency or ("KRW" if m.amount is not None else None)
                        vals += [cur, m.amount]
                    else:
                        vals.append(_fmt_cell(cell))
            for j, v in enumerate(vals):
                c = ws.cell(r, j + 1, v)
                if j in money_out_idx:
                    _money_fmt(c)
            # 조서번호(A)는 빈 입력칸 — 0_요약 시트에서 기관별로 관리.
            r += 1
    ws.freeze_panes = "A3"
    _autosize(ws, max_width=46, min_row=2)  # 1행 머리말 문장은 폭 계산 제외


def _present_sections(category: str, records: list[ConfirmationRecord]) -> list:
    """해당 카테고리에서 데이터(표)가 1건이라도 있는 섹션 스펙 목록."""
    specs = load_section_spec().get(category, [])
    present = []
    for spec in specs:
        if any((rec.section_tables.get(spec.id) and rec.section_tables[spec.id].rows)
               for rec in records):
            present.append(spec)
    return present


def _build_status_sheet(ws, records, config) -> None:
    """데이터 섹션이 없는 카테고리(예: 비거래 서울보증)용 취합현황 시트."""
    for j, c in enumerate(["조서번호", "금융기관명", "사업자번호", "조회기준일", "취합 현황"], start=1):
        ws.cell(1, j, c)
    _style_header(ws, 1, 5)
    r = 2
    for rec in sorted(records, key=lambda x: (x.header.institution_name, x.header.business_no)):
        h = rec.header
        ws.cell(r, 2, h.institution_name); ws.cell(r, 3, h.business_no)
        ws.cell(r, 4, h.confirmation_date); ws.cell(r, 5, config.status_label(rec.status))
        r += 1
    ws.freeze_panes = "A2"
    _autosize(ws)


def build_category_workbook(category, records, config, client) -> tuple[Workbook, list[str]]:
    wb = Workbook()
    wb.remove(wb.active)
    sheet_names: list[str] = []
    shorts = _SECTION_SHORT.get(category, {})
    present = _present_sections(category, records)
    # 데이터 섹션 시트
    for spec in present:
        name = _safe_sheet_name(f"{spec.id}_{shorts.get(spec.id, spec.id)}")
        render_section_sheet(wb.create_sheet(name), spec.id, spec, records, config, client)
        sheet_names.append(name)
    # 0_요약 시트(라이브 수식)를 맨 앞에 삽입
    if any(_money_fields(spec) for spec in present):
        ws = wb.create_sheet("0_요약", index=0)
        build_credit_summary_sheet(ws, category, records, present, config)
        sheet_names.insert(0, "0_요약")
    if not sheet_names:  # 데이터 섹션 없음 → 취합현황만
        _build_status_sheet(wb.create_sheet("취합현황"), records, config)
        sheet_names.append("취합현황")
    return wb, sheet_names


def write_outputs(records, config, client: str, out_dir: Path, pbc_result=None) -> list[Path]:
    """용도별 다중 파일 저장. 저장된 경로 목록 반환."""
    out_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    file_sheets: dict[str, list[str]] = {}

    by_cat: dict[str, list[ConfirmationRecord]] = {}
    for rec in records:
        by_cat.setdefault(rec.header.institution_category, []).append(rec)

    # 1) 카테고리별 파일 (1섹션 = 1시트)
    for category, tag in _CATEGORY_FILE.items():
        recs = by_cat.get(category)
        if not recs:
            continue
        wb, sheets = build_category_workbook(category, recs, config, client)
        path = out_dir / f"금융기관조회서_{tag}_{client}.xlsx"
        wb.save(path)
        written.append(path)
        file_sheets[path.name] = sheets

    # 2) PBC대사 파일 (분석 시트)
    recon_rows = build_recon_rows(records, load_account_mapping())
    if recon_rows:
        wb = Workbook(); wb.remove(wb.active)
        build_recon_sheet(wb.create_sheet("대사"), recon_rows, client)
        _write_summary(wb.create_sheet("계정과목별"), recon_rows, "category", "계정과목", client)
        _write_summary(wb.create_sheet("금융기관별"), recon_rows, "institution", "금융기관", client)
        _write_summary(wb.create_sheet("통화별"), recon_rows, "currency", "통화", client)
        if pbc_result is not None:
            build_pbc_recon_sheet(wb.create_sheet("PBC대사"), pbc_result, client)
        path = out_dir / f"금융기관조회서_PBC대사_{client}.xlsx"
        wb.save(path); written.append(path)
        file_sheets[path.name] = wb.sheetnames

    # 요약은 별도 파일이 아니라 각 카테고리 파일의 '0_요약' 시트(라이브 수식)로 들어간다.
    return written
