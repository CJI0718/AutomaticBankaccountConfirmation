"""대사·계정과목별·통화별·PBC대사 시트 빌더 (output.py 가 사용).

각 함수는 워크시트(ws)를 받아 채운다. 다중 파일 구성은 afc/output.py 참고.
"""
from __future__ import annotations

from collections.abc import Iterable

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from afc.extract import ConfirmationRecord
from afc.institutions import InstitutionConfig
from afc.reconciliation import ReconRow

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SUBTOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
_TOTAL_FILL = PatternFill("solid", fgColor="FCE4D6")
_PBC_FILL = PatternFill("solid", fgColor="FFF2CC")  # 감사인 입력 칸 (회사제시/환율)
_RECON_COLS = [
    "계정분류", "구분", "금융기관", "계좌번호", "금융상품종류", "통화",
    "환산전 금액", "적용환율", "환산후 금액(원화)",
    "이자율", "만기", "사용제한",
    "[회사제시] 환산후(원화)", "차이(원화)", "분류근거",
]
_FX_COL, _PRE_COL, _POST_COL = 8, 7, 9          # 적용환율 / 환산전 / 환산후
_PBC_POST_COL, _DIFF_COL = 13, 14
_KRW = {"KRW", "WON", None}
_SUMMARY_COLS = ["조서 번호", "금융기관명", "사업자등록번호", "취합 현황"]
_DEPOSIT_COLS = [
    "조회대상 회사", "사업자번호", "금융기관명", "조회기준일",
    "금융상품의 종류", "계좌번호", "통화", "금액",
    "연이자율", "최종이자지급일", "만기일", "인출제한 등",
]


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _display_width(value) -> int:
    """한글·전각은 라틴 대비 폭이 넓어 약 1.8배로 환산."""
    s = str(value)
    return sum(2 if ord(ch) > 0x2E7F else 1 for ch in s)


def _autosize(ws, max_width: int = 50, min_row: int = 1) -> None:
    """열 너비 자동. min_row 이전 행(예: 1행의 긴 섹션 머리말)은 폭 계산에서 제외해
    열자마자 표가 보기 좋게 한다(머리말은 빈 오른쪽 칸으로 자연히 넘쳐 보임)."""
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        widths = [
            _display_width(c.value) for c in col
            if c.value is not None and c.row >= min_row
            and not (isinstance(c.value, str) and c.value.startswith("="))  # 수식은 결과가 짧음
        ]
        width = max(widths, default=8)
        ws.column_dimensions[letter].width = min(max_width, max(9, width + 2))


def _summary_rows(records: Iterable[ConfirmationRecord], config: InstitutionConfig):
    rows = []
    for rec in records:
        rows.append(
            (
                rec.header.institution_name,
                rec.header.business_no,
                config.status_label(rec.status),
            )
        )
    # cck ordering: institution name (Korean), then business number.
    rows.sort(key=lambda r: (r[0], r[1]))
    return rows


def _money_fmt(cell) -> None:
    cell.number_format = "#,##0.00"


def build_recon_sheet(ws, rows: list[ReconRow], client: str) -> None:
    """감사조서 4200 '조회서대사' 축으로 정리한 시트.

    계정분류 → 통화 순으로 묶고, (계정분류·통화) 소계와 계정분류별 환산후(원화)
    roll-up, 전체 총계를 SUM 수식으로 넣는다. 적용환율·회사제시 칸은 노란색으로
    비워 두어 감사인이 채우면 환산후/차이가 자동 계산된다.
    """
    ws.cell(1, 1, f"프로젝트명: {client}").font = Font(bold=True)
    ws.cell(2, 1, "금융기관조회서 ↔ 회사제시(PBC) 대사").font = Font(bold=True, size=12)
    hdr = 4
    for j, name in enumerate(_RECON_COLS, start=1):
        ws.cell(hdr, j, name)
    _style_header(ws, hdr, len(_RECON_COLS))

    r = hdr + 1
    post_by_side: dict[str, list[int]] = {"자산": [], "부채": []}

    # 계정분류 → 통화 그룹 경계에서 소계를 삽입.
    from itertools import groupby
    for category, cat_iter in groupby(rows, key=lambda x: x.category):
        cat_rows = list(cat_iter)
        for currency, cur_iter in groupby(cat_rows, key=lambda x: x.currency):
            cur_rows = list(cur_iter)
            first = r
            for row in cur_rows:
                is_krw = row.currency in _KRW
                ws.cell(r, 1, row.category)
                ws.cell(r, 2, row.side)
                ws.cell(r, 3, row.institution)
                ws.cell(r, 4, row.account_no)
                ws.cell(r, 5, row.product)
                ws.cell(r, 6, row.currency)
                _money_fmt(ws.cell(r, _PRE_COL, row.amount))
                fx = ws.cell(r, _FX_COL, 1 if is_krw else None)
                if not is_krw:
                    fx.fill = _PBC_FILL  # 외화 환율은 감사인 입력
                post = ws.cell(r, _POST_COL, f"=G{r}*H{r}")
                _money_fmt(post)
                ws.cell(r, 10, row.interest_rate)
                ws.cell(r, 11, row.maturity)
                ws.cell(r, 12, row.restrictions)
                ws.cell(r, _PBC_POST_COL).fill = _PBC_FILL          # 회사제시 입력칸
                _money_fmt(ws.cell(r, _DIFF_COL, f"=I{r}-M{r}"))    # 차이 = 조회서 - PBC
                ws.cell(r, 15, row.basis)
                post_by_side.setdefault(row.side, []).append(r)
                r += 1
            # (계정분류·통화) 소계 — 환산전 합계
            ws.cell(r, 5, f"  ▷ {category} / {currency or '원화'} 소계")
            _money_fmt(ws.cell(r, _PRE_COL, f"=SUM(G{first}:G{r-1})"))
            _money_fmt(ws.cell(r, _POST_COL, f"=SUM(I{first}:I{r-1})"))
            for c in range(1, len(_RECON_COLS) + 1):
                ws.cell(r, c).fill = _SUBTOTAL_FILL
            ws.cell(r, 5).font = Font(italic=True)
            r += 1
    # 자산/부채 총계 (환산후 원화) — 절대 섞지 않는다.
    r += 1
    for side in ("자산", "부채"):
        idxs = post_by_side.get(side) or []
        if not idxs:
            continue
        ws.cell(r, 5, f"■ {side} 총계 (환산후 원화)").font = Font(bold=True)
        _money_fmt(ws.cell(r, _POST_COL, "=" + "+".join(f"I{x}" for x in idxs)))
        for c in range(1, len(_RECON_COLS) + 1):
            ws.cell(r, c).fill = _TOTAL_FILL
        r += 1

    ws.freeze_panes = "A5"
    _autosize(ws)


def build_pbc_recon_sheet(ws, result, client: str) -> None:
    """조회서 ↔ 회사제시(PBC) 계좌 단위 대사 + 양방향 완전성 예외.

    감사조서 4200 절차: 회사명세→조회서, 조회서→회사명세 두 방향을 모두 확인한다.
    """
    cols = ["계좌번호", "금융기관", "계정분류", "통화", "조회서 금액", "PBC 금액", "차이", "상태"]
    ws.cell(1, 1, f"프로젝트명: {client}").font = Font(bold=True)
    ws.cell(2, 1, "조회서 ↔ 회사제시(PBC) 대사").font = Font(bold=True, size=12)
    s = result.summary()
    ws.cell(3, 1, f"일치 {s['matched']-s['diff']} · 차이 {s['diff']} · "
                  f"조회서에만(명세누락) {s['confirm_only']} · 명세에만(미회수) {s['pbc_only']}")

    hdr = 5
    for j, name in enumerate(cols, start=1):
        ws.cell(hdr, j, name)
    _style_header(ws, hdr, len(cols))
    r = hdr + 1

    ws.cell(r, 1, "■ 계좌 대사 결과").font = Font(bold=True)
    r += 1
    for m in result.matched:
        ws.cell(r, 1, m.account_key)
        ws.cell(r, 2, m.institution)
        ws.cell(r, 3, m.category)
        ws.cell(r, 4, ", ".join(m.currencies))
        _money_fmt(ws.cell(r, 5, m.confirm_amount))
        _money_fmt(ws.cell(r, 6, m.pbc_amount))
        _money_fmt(ws.cell(r, 7, m.diff))
        st = ws.cell(r, 8, m.status)
        if m.status == "차이":
            for c in range(1, len(cols) + 1):
                ws.cell(r, c).fill = _TOTAL_FILL
            st.font = Font(bold=True, color="C00000")
        r += 1

    if result.confirm_only:
        r += 1
        ws.cell(r, 1, "■ 조회서에만 존재 — 회사명세 누락 (조회서→명세 완전성)").font = Font(bold=True, color="C00000")
        r += 1
        for key, inst, cat, amt in result.confirm_only:
            ws.cell(r, 1, key); ws.cell(r, 2, inst); ws.cell(r, 3, cat)
            _money_fmt(ws.cell(r, 5, amt))
            ws.cell(r, 8, "명세누락")
            for c in range(1, len(cols) + 1):
                ws.cell(r, c).fill = _SUBTOTAL_FILL
            r += 1

    if result.pbc_only:
        r += 1
        ws.cell(r, 1, "■ 회사명세에만 존재 — 조회서 미회수 (명세→조회서 완전성)").font = Font(bold=True, color="C00000")
        r += 1
        for p in result.pbc_only:
            ws.cell(r, 1, p.account_key); ws.cell(r, 2, p.institution)
            ws.cell(r, 4, p.currency)
            _money_fmt(ws.cell(r, 6, p.amount))
            ws.cell(r, 8, "미회수")
            for c in range(1, len(cols) + 1):
                ws.cell(r, c).fill = _PBC_FILL
            r += 1

    ws.freeze_panes = "A6"
    _autosize(ws)


def _aggregate(rows: list[ReconRow], *keys):
    """(key tuple) -> [건수, 환산전합계]. 통화는 항상 키에 포함해 외화 혼합 합산을 방지."""
    from collections import defaultdict
    agg: dict[tuple, list] = defaultdict(lambda: [0, 0.0])
    for r in rows:
        k = tuple((r.currency or "KRW") if key == "currency" else getattr(r, key) for key in keys)
        slot = agg[k]
        slot[0] += 1
        slot[1] += r.amount or 0.0
    return agg


_CUR_ORDER = ["KRW", "USD", "EUR", "JPY", "CNY", "HKD"]


def _cat_rank(cat: str) -> int:
    return ReconRow._ORDER.index(cat) if cat in ReconRow._ORDER else 99


def _write_summary(ws, rows: list[ReconRow], primary: str, primary_label: str, client: str) -> None:
    """구분(자산/부채) → primary(계정과목/금융기관/통화) → 통화 roll-up.
    자산과 부채, 통화는 절대 섞어 합산하지 않는다."""
    is_currency_primary = primary == "currency"
    cols = (["구분", "통화", "건수", "환산전 합계"] if is_currency_primary
            else ["구분", primary_label, "통화", "건수", "환산전 합계", "환산후(원화)*"])
    ws.cell(1, 1, f"프로젝트명: {client}").font = Font(bold=True)
    ws.cell(2, 1, f"{primary_label} 정리 (자산/부채 구분)").font = Font(bold=True, size=12)
    if not is_currency_primary:
        ws.cell(3, 1, "* 환산후(원화)는 KRW만 표시 — 외화는 '대사' 시트에서 환율 입력 후 합산").font = Font(size=9, italic=True)
    hdr = 4
    for j, name in enumerate(cols, start=1):
        ws.cell(hdr, j, name)
    _style_header(ws, hdr, len(cols))
    r = hdr + 1

    if is_currency_primary:
        agg = _aggregate(rows, "side", "currency")
        for key in sorted(agg, key=lambda x: (x[0] != "자산", _CUR_ORDER.index(x[1]) if x[1] in _CUR_ORDER else 99, x[1])):
            side, cur = key
            cnt, total = agg[key]
            ws.cell(r, 1, side); ws.cell(r, 2, cur); ws.cell(r, 3, cnt)
            _money_fmt(ws.cell(r, 4, round(total, 2)))
            r += 1
        _autosize(ws)
        return

    agg = _aggregate(rows, "side", primary, "currency")
    groups = sorted({(k[0], k[1]) for k in agg},
                    key=lambda x: (x[0] != "자산", _cat_rank(x[1]) if primary == "category" else 0, x[1]))
    for side, p in groups:
        first = r
        for cur in sorted({k[2] for k in agg if k[0] == side and k[1] == p},
                          key=lambda c: (_CUR_ORDER.index(c) if c in _CUR_ORDER else 99, c)):
            cnt, total = agg[(side, p, cur)]
            ws.cell(r, 1, side); ws.cell(r, 2, p); ws.cell(r, 3, cur); ws.cell(r, 4, cnt)
            _money_fmt(ws.cell(r, 5, round(total, 2)))
            if cur in _KRW:
                _money_fmt(ws.cell(r, 6, round(total, 2)))
            r += 1
        if r - first > 1:  # 소계: 건수(D) 합 + 환산후 원화(F, KRW만) 합
            ws.cell(r, 2, f"  ▷ {p} 소계").font = Font(italic=True)
            ws.cell(r, 4, f"=SUM(D{first}:D{r-1})")
            _money_fmt(ws.cell(r, 6, f"=SUM(F{first}:F{r-1})"))
            for c in range(1, len(cols) + 1):
                ws.cell(r, c).fill = _SUBTOTAL_FILL
            r += 1
    _autosize(ws)
