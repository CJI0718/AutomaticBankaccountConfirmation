from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class CellDiff:
    sheet: str
    cell: str
    expected: object
    actual: object


def normalize_empty(value: object) -> object:
    return None if value == "" else value


def diff_xlsx(produced_path: Path, expected_path: Path) -> list[CellDiff]:
    produced = load_workbook(produced_path, data_only=True)
    expected = load_workbook(expected_path, data_only=True)
    diffs: list[CellDiff] = []

    for sheet_name in expected.sheetnames:
        if sheet_name not in produced.sheetnames:
            diffs.append(CellDiff(sheet_name, "<sheet>", "present", "missing"))
            continue
        ws_expected = expected[sheet_name]
        ws_produced = produced[sheet_name]
        max_row = max(ws_expected.max_row, ws_produced.max_row)
        max_col = max(ws_expected.max_column, ws_produced.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                expected_value = normalize_empty(ws_expected.cell(row, col).value)
                actual_value = normalize_empty(ws_produced.cell(row, col).value)
                if expected_value != actual_value:
                    diffs.append(
                        CellDiff(
                            sheet=sheet_name,
                            cell=ws_expected.cell(row, col).coordinate,
                            expected=expected_value,
                            actual=actual_value,
                        )
                    )
    return diffs


def main() -> None:
    parser = argparse.ArgumentParser(description="Cell-level xlsx diff against a cck benchmark workbook.")
    parser.add_argument("produced", type=Path)
    parser.add_argument("expected", type=Path)
    parser.add_argument("--limit", type=int, default=100)
    args = parser.parse_args()

    diffs = diff_xlsx(args.produced, args.expected)
    print(json.dumps({"diff_count": len(diffs)}, ensure_ascii=False))
    for diff in diffs[: args.limit]:
        print(json.dumps(asdict(diff), ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
