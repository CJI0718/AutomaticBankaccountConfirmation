"""PBC 로더 + 조회서↔PBC 대사 엔진 테스트."""
from __future__ import annotations

from openpyxl import Workbook

from afc.pbc import load_pbc, normalize_account, reconcile
from afc.reconciliation import ReconRow


def _recon(account, currency, amount, institution="신한은행", category="현금및현금성자산"):
    return ReconRow(
        category=category, side="자산", institution=institution, company="가나㈜",
        business_no="123-45-67890", account_no=account, product="보통예금",
        currency=currency, amount=amount, interest_rate=None, maturity="00000000",
        restrictions=None, basis="",
    )


def test_normalize_account():
    assert normalize_account("087-17-022205") == "08717022205"
    assert normalize_account("100-035-030829") == "100035030829"
    assert normalize_account(None) == ""


def _write_pbc(path, rows, header=("금융기관", "적요", "계좌번호", "통화", "금액", "비고")):
    wb = Workbook()
    ws = wb.active
    ws.title = "보통예금"
    ws.cell(2, 1, "보 통 예 금 명 세 서")
    for j, h in enumerate(header, start=1):
        ws.cell(5, j, h)
    r = 6
    for row in rows:
        for j, v in enumerate(row, start=1):
            ws.cell(r, j, v)
        r += 1
    ws.cell(r, 1, "합 계")  # 집계 행은 스킵돼야 함
    ws.cell(r, 5, 999)
    wb.save(path)


def test_load_pbc_skips_total_and_normalizes(tmp_path):
    p = tmp_path / "pbc.xlsx"
    _write_pbc(p, [
        ("신한은행", "보통예금", "100-035-030829", "KRW", 1000, ""),
        ("우리은행", "정기예금", "1005-202-813898", "KRW", 2000, ""),
    ])
    rows = load_pbc(p)
    assert len(rows) == 2  # 합계 행 제외
    by_key = {r.account_key: r for r in rows}
    assert "100035030829" in by_key
    assert by_key["100035030829"].amount == 1000


def test_reconcile_match_diff_and_both_completeness(tmp_path):
    p = tmp_path / "pbc.xlsx"
    _write_pbc(p, [
        ("신한은행", "보통예금", "100-000-111", "KRW", 1000, ""),   # 일치
        ("신한은행", "보통예금", "100-000-222", "KRW", 1500, ""),   # 차이 (조회서 2000)
        ("우리은행", "정기예금", "100-000-999", "KRW", 700, ""),    # 조회서에 없음 -> pbc_only
    ])
    confirm = [
        _recon("100000111", "KRW", 1000),                          # 일치
        _recon("100000222", "KRW", 2000),                          # 차이
        _recon("100000333", "KRW", 50, institution="국민은행"),    # PBC에 없음 -> confirm_only
    ]
    result = reconcile(confirm, load_pbc(p))
    s = result.summary()
    assert s["matched"] == 2
    assert s["diff"] == 1
    assert s["confirm_only"] == 1
    assert s["pbc_only"] == 1

    diff_row = next(m for m in result.matched if m.status == "차이")
    assert diff_row.account_key == "100000222"
    assert diff_row.diff == 500.0


def test_currency_aware_matching(tmp_path):
    """같은 계좌라도 통화가 다르면 별도 대사."""
    p = tmp_path / "pbc.xlsx"
    _write_pbc(p, [
        ("신한은행", "외화예금", "100-000-500", "USD", 100, ""),
        ("신한은행", "외화예금", "100-000-500", "EUR", 200, ""),
    ])
    confirm = [_recon("100000500", "USD", 100), _recon("100000500", "EUR", 999)]
    result = reconcile(confirm, load_pbc(p))
    s = result.summary()
    assert s["matched"] == 2 and s["diff"] == 1
